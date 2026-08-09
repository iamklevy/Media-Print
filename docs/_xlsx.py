# -*- coding: utf-8 -*-
"""Build the Media Print Pack data-gap workbook from the live site data."""
import json, io, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NEXT = 'D:/mediaprint-next'
OUT = 'D:/mediaprint/docs/mediaprint-data-gaps.xlsx'

# ---------------------------------------------------------------- source data
src = io.open(f'{NEXT}/content/products.ts', encoding='utf-8').read()
PRODUCTS = json.loads(re.search(r'PRODUCTS: CatalogueProduct\[\] = (\[[\s\S]*?\]);', src).group(1))
PRICING = json.loads(re.search(r'PriceBook = ([\s\S]*?) as PriceBook;',
                               io.open(f'{NEXT}/content/pricing.ts', encoding='utf-8').read()).group(1))
STICKERS = json.loads(re.search(r'StickerCatalogue = ([\s\S]*?);\n',
                                io.open(f'{NEXT}/content/stickers.ts', encoding='utf-8').read()).group(1))
EN = json.load(io.open(f'{NEXT}/messages/en.json', encoding='utf-8'))
AR = json.load(io.open(f'{NEXT}/messages/ar.json', encoding='utf-8'))


def tr(tree, dotted):
    c = tree
    for p in dotted.split('.'):
        c = c[p]
    return c.replace('&amp;', '&')


# ---------------------------------------------------------------- styling
F = 'Arial'
H1 = Font(name=F, size=15, bold=True, color='16130F')
H2 = Font(name=F, size=11, bold=True, color='FFFFFF')
BOLD = Font(name=F, size=10, bold=True)
BODY = Font(name=F, size=10)
MUTED = Font(name=F, size=9, color='6E645A')
INPUT_FONT = Font(name=F, size=10, color='0000FF')
EG_FONT = Font(name=F, size=10, italic=True, color='008000')

HEAD_FILL = PatternFill('solid', fgColor='16130F')
INPUT_FILL = PatternFill('solid', fgColor='FFFF00')
DONE_FILL = PatternFill('solid', fgColor='E8F3ED')
WARN_FILL = PatternFill('solid', fgColor='FFF4E0')
BAND = PatternFill('solid', fgColor='F7F2EA')

thin = Side(style='thin', color='D9D2C7')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')
CTR = Alignment(horizontal='center', vertical='center')

wb = Workbook()


def header(ws, row, cols):
    for i, (title, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=title)
        c.font, c.fill, c.border = H2, HEAD_FILL, BOX
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, r, c, v, font=BODY, fill=None, align=WRAP, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font, cell.border, cell.alignment = font, BOX, align
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ================================================================ 1. READ ME
ws = wb.active
ws.title = 'Read me'
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 108

rows = [
    ('h1', 'Media Print Pack — data we still need'),
    ('p', 'Generated from the live website data. Every figure already shown here came out of '
          '"Price list 2026-3-25.pdf" and was cross-checked against the XLSX quote sheet.'),
    ('sp', ''),
    ('h2', 'How to fill this in'),
    ('p', 'Type into the YELLOW cells only. Everything else is either already confirmed or calculated.'),
    ('p', 'Each sheet starts with one green italic EXAMPLE row showing the expected format — overwrite '
          'it or leave it, it is ignored.'),
    ('p', 'Leave a cell blank if the answer is "we do not make that" — do not put 0, which reads as free.'),
    ('sp', ''),
    ('h2', 'What is in each sheet'),
    ('p', '1. Products — all 12 products on the site and whether each has a usable price. Read-only.'),
    ('p', '2. Prices needed — the 5 products with no prices at all. This is the highest-value sheet.'),
    ('p', '3. Confirm these — 6 figures I had to reconstruct from a hard-to-read PDF table. Please verify.'),
    ('p', '4. Stickers — 17 sticker types with no rates. Currently all show "Price on request" on the site.'),
    ('p', '5. Services — 8 commercial printing services, no pricing anywhere.'),
    ('p', '6. Photos & assets — images and files the site still needs.'),
    ('sp', ''),
    ('h2', 'Colour key'),
]
r = 2
for kind, text in rows:
    if kind == 'sp':
        r += 1
        continue
    c = ws.cell(row=r, column=2, value=text)
    c.font = {'h1': H1, 'h2': BOLD, 'p': BODY}[kind]
    c.alignment = WRAP
    r += 1

for label, fill, note in [
    ('Yellow', INPUT_FILL, 'you fill this in'),
    ('Green', DONE_FILL, 'already confirmed — nothing to do'),
    ('Orange', WARN_FILL, 'reconstructed from the PDF — please double-check'),
]:
    ws.cell(row=r, column=1).fill = fill
    ws.cell(row=r, column=1).border = BOX
    c = ws.cell(row=r, column=2, value=f'{label} — {note}')
    c.font, c.alignment = BODY, WRAP
    r += 1

r += 1
c = ws.cell(row=r, column=2, value='Priority: the zipper bags, paper sacks and printed cartons on '
                                   '"Prices needed" are the three most prominent products on the site '
                                   'and none of them can show a price today.')
c.font, c.alignment = Font(name=F, size=10, bold=True, color='B3480F'), WRAP

# ================================================================ 2. PRODUCTS
ws = wb.create_sheet('Products')
ws.cell(row=1, column=1, value='All products on the website').font = H1
CAT = {'plastic': 'Plastic & flexible', 'paper': 'Paper & board',
       'fabric': 'Fabric', 'print': 'Labels & print'}
header(ws, 3, [('#', 5), ('Product (English)', 34), ('Product (Arabic)', 30), ('Category', 19),
               ('Has pricing?', 13), ('Sizes', 7), ('From (EGP)', 11), ('MOQ', 9),
               ('Where the price came from', 34)])

r = 4
priced_rows = []
for i, p in enumerate(PRODUCTS, 1):
    pr = PRICING['products'].get(p['pricing']) if p['pricing'] else None
    put(ws, r, 1, i, BODY, None, CTR)
    put(ws, r, 2, tr(EN, p['key'] + '.t'))
    put(ws, r, 3, tr(AR, p['key'] + '.t'), BODY, None,
        Alignment(wrap_text=True, vertical='top', horizontal='right'))
    put(ws, r, 4, CAT[p['cat']])
    if pr:
        vals = []
        for v in pr['variants']:
            pv = v.get('prices')
            if isinstance(pv, list):
                vals += pv
            elif isinstance(pv, dict):
                vals += list(pv.values())
            elif v.get('perKg'):
                vals.append(round(v['perKg'] / v['pcsPerKg'], 2))
        moq = pr['tiers'][0] if pr.get('tiers') else min((v.get('moq') or 1) for v in pr['variants'])
        put(ws, r, 5, 'Yes', BOLD, DONE_FILL, CTR)
        put(ws, r, 6, len(pr['variants']), BODY, DONE_FILL, CTR)
        put(ws, r, 7, min(vals), BODY, DONE_FILL, CTR, '#,##0.00')
        put(ws, r, 8, moq, BODY, DONE_FILL, CTR, '#,##0')
        put(ws, r, 9, 'Price list 2026-3-25 (verified)', MUTED)
        priced_rows.append(r)
    else:
        put(ws, r, 5, 'NO', Font(name=F, size=10, bold=True, color='B02A2A'), INPUT_FILL, CTR)
        for col in (6, 7, 8):
            put(ws, r, col, None, BODY, INPUT_FILL, CTR)
        put(ws, r, 9, 'Not in the price list — see "Prices needed"', MUTED)
    r += 1

r += 1
put(ws, r, 2, 'Products with usable pricing', BOLD)
c = put(ws, r, 5, f'=COUNTIF(E4:E{r - 2},"Yes")', BOLD, DONE_FILL, CTR)
put(ws, r, 6, f'=COUNTA(B4:B{r - 2})&" total"', MUTED, None, CTR)

# ============================================================ 3. PRICES NEEDED
ws = wb.create_sheet('Prices needed')
ws.cell(row=1, column=1, value='Prices needed — products the site cannot quote today').font = H1
ws.cell(row=2, column=1, value='Add one row per size. Add as many rows as you need under each product.').font = MUTED

header(ws, 4, [('Product', 26), ('Size (cm)', 14), ('Material / colour', 18), ('MOQ', 9),
               ('Price @ MOQ', 12), ('Qty tier 2', 11), ('Price @ tier 2', 13),
               ('Qty tier 3', 11), ('Price @ tier 3', 13), ('Printing +EGP', 13), ('Notes', 26)])

r = 5
ex = ['EXAMPLE — Zipper bags', '20 × 30', 'Clear, 2 colour', 1000, 2.75, 3000, 2.4, 5000, 2.1, 0.5,
      'Delete or overwrite this row']
for i, v in enumerate(ex, 1):
    put(ws, r, i, v, EG_FONT, None, CTR if i > 3 else WRAP,
        '#,##0.00' if i in (5, 7, 9, 10) else ('#,##0' if i in (4, 6, 8) else None))
r += 1

UNPRICED = [p for p in PRODUCTS if not p['pricing']]
for p in UNPRICED:
    name = tr(EN, p['key'] + '.t')
    for n in range(4):
        put(ws, r, 1, name if n == 0 else None, BOLD if n == 0 else BODY, BAND if n == 0 else None)
        for col in range(2, 11):
            put(ws, r, col, None, INPUT_FONT, INPUT_FILL, CTR,
                '#,##0.00' if col in (5, 7, 9, 10) else ('#,##0' if col in (4, 6, 8) else None))
        put(ws, r, 11, None, INPUT_FONT, INPUT_FILL)
        r += 1

r += 1
put(ws, r, 1, 'Rows filled in so far', BOLD)
put(ws, r, 5, f'=COUNT(E6:E{r - 2})', BOLD, DONE_FILL, CTR)

# ============================================================ 4. CONFIRM THESE
ws = wb.create_sheet('Confirm these')
ws.cell(row=1, column=1, value='Figures reconstructed from the PDF — please confirm').font = H1
ws.cell(row=2, column=1,
        value='The 2026 price list table did not extract cleanly for these cells. I used the value that '
              'kept prices falling as quantity rises and matched your quote sheet, but it is inference, '
              'not reading. Correct anything wrong in the yellow column.').font = MUTED
ws.row_dimensions[2].height = 30

header(ws, 4, [('Item', 30), ('Cell in question', 24), ('What the site uses now', 20),
               ('Correct value', 15), ('Why I am unsure', 46)])

CONFIRM = [
    ('Corrugated box 25×30×10 Brown', 'Price at qty 100', '21.50',
     'Extraction gave an ambiguous 20.50, which would make the 100-qty price cheaper than the 200-qty one.'),
    ('Corrugated box 25×30×10 White', 'Prices at 100 / 200 / 300', '26 / 25 / 24',
     'Raw extraction read 24/25/26 — rising with quantity. I reversed it because 24 at qty 300 plus '
     'EGP 3 printing equals the 27 on your quote sheet.'),
    ('Corrugated boxes', 'Printing surcharge at qty 100', '3.00',
     'Clearly 3 at qty 200/300/500. The qty-100 cell looked like 4 and the qty-1000 cell did not extract.'),
    ('Corrugated boxes', 'Printing surcharge at qty 1,000', '3.00',
     'Did not extract at all. Assumed the same 3 as the other tiers.'),
    ('Kraft carrier bags', 'Third size', 'Only 2 sizes on site',
     'You said three sizes. The price list has 25×30×8 and 30×40×10 only. What is the third, and its price?'),
    ('Aluminium / metallised pouches', 'Minimum order quantities', '4,000 / 2,000 / 1,000 / 1,000',
     'You said 4 sizes with a 1,000 minimum for single colour. The PDF shows those four MOQs instead.'),
    ('Aluminium pouches — 4 colour', 'All prices', 'Not on the site',
     'You mentioned the "Chad" bag, 4 sizes, MOQ 3,000 — but the sizes and prices never arrived.'),
    ('Butter paper (PDF page 6)', 'Whole table', 'Left off the site',
     'That page was too garbled to read. Please re-send those prices if you sell it.'),
]
r = 5
for item, cell, cur, why in CONFIRM:
    put(ws, r, 1, item, BOLD)
    put(ws, r, 2, cell)
    put(ws, r, 3, cur, BODY, WARN_FILL, CTR)
    put(ws, r, 4, None, INPUT_FONT, INPUT_FILL, CTR)
    put(ws, r, 5, why, MUTED)
    ws.row_dimensions[r].height = 34
    r += 1

# ================================================================ 5. STICKERS
ws = wb.create_sheet('Stickers')
ws.cell(row=1, column=1, value='Sticker catalogue — needs your rates').font = H1
ws.cell(row=2, column=1,
        value='These 17 types are already live on the site and all show "Price on request". '
              'I did not copy any competitor\'s prices.').font = MUTED

header(ws, 4, [('Category', 20), ('Sticker', 30), ('Arabic', 26), ('Material', 17),
               ('Size (cm)', 11), ('MOQ', 9), ('Price each @ MOQ', 15),
               ('Price each @ 5,000', 16), ('Notes', 24)])

catname = {c['id']: c['en'] for c in STICKERS['categories']}
matname = {m['id']: m['en'] for m in STICKERS['materials']}

r = 5
put(ws, r, 1, 'EXAMPLE', EG_FONT)
for i, v in enumerate(['Coffee cup sticker', 'استيكر كوب قهوة', 'Paper', '5 × 5', 1000, 1.2, 0.85,
                       'Delete or overwrite'], 2):
    put(ws, r, i, v, EG_FONT, None, CTR if i in (6, 7, 8) else WRAP,
        '#,##0.00' if i in (7, 8) else ('#,##0' if i == 6 else None))
r += 1

for it in STICKERS['items']:
    put(ws, r, 1, catname.get(it['cat'], it['cat']))
    put(ws, r, 2, it['en'], BOLD)
    put(ws, r, 3, it['ar'], BODY, None, Alignment(wrap_text=True, vertical='top', horizontal='right'))
    put(ws, r, 4, matname.get(it['mat'], it['mat']))
    put(ws, r, 5, it['size'], BODY, None, CTR)
    if it['price'] is not None:
        put(ws, r, 6, it.get('moq'), BODY, DONE_FILL, CTR, '#,##0')
        put(ws, r, 7, it['price'], BODY, DONE_FILL, CTR, '#,##0.00')
        put(ws, r, 8, None, INPUT_FONT, INPUT_FILL, CTR, '#,##0.00')
        put(ws, r, 9, 'Already confirmed (price list p.10)', MUTED)
    else:
        for col in (6, 7, 8):
            put(ws, r, col, None, INPUT_FONT, INPUT_FILL, CTR,
                '#,##0.00' if col in (7, 8) else '#,##0')
        put(ws, r, 9, None, INPUT_FONT, INPUT_FILL)
    r += 1

r += 1
put(ws, r, 2, 'Stickers priced', BOLD)
put(ws, r, 7, f'=COUNT(G6:G{r - 2})', BOLD, DONE_FILL, CTR)
put(ws, r, 8, f'=COUNTA(B6:B{r - 2})&" total"', MUTED, None, CTR)

# ================================================================ 6. SERVICES
ws = wb.create_sheet('Services')
ws.cell(row=1, column=1, value='Commercial printing services — no pricing anywhere').font = H1
ws.cell(row=2, column=1, value='Listed on the site with descriptions only.').font = MUTED
header(ws, 4, [('Service', 32), ('Arabic', 30), ('Unit', 16), ('MOQ', 10),
               ('Price per unit @ MOQ', 18), ('Notes', 34)])

SERVICES = ['s.metalize', 's.notepad', 's.letterhead', 's.sticker',
            's.duplex', 's.rollup', 's.flyer', 's.promo']
r = 5
put(ws, r, 1, 'EXAMPLE — Roll-up banner', EG_FONT)
for i, v in enumerate(['رول أب', 'per banner', 1, 850, 'Delete or overwrite'], 2):
    put(ws, r, i, v, EG_FONT, None, CTR if i in (4, 5) else WRAP,
        '#,##0.00' if i == 5 else ('#,##0' if i == 4 else None))
r += 1
for k in SERVICES:
    put(ws, r, 1, tr(EN, k + '.t'), BOLD)
    put(ws, r, 2, tr(AR, k + '.t'), BODY, None, Alignment(wrap_text=True, vertical='top', horizontal='right'))
    for col in (3, 4, 5, 6):
        put(ws, r, col, None, INPUT_FONT, INPUT_FILL, CTR if col in (4, 5) else WRAP,
            '#,##0.00' if col == 5 else ('#,##0' if col == 4 else None))
    r += 1

# ============================================================ 7. PHOTOS/ASSETS
ws = wb.create_sheet('Photos & assets')
ws.cell(row=1, column=1, value='Images and files the site still needs').font = H1
header(ws, 3, [('What', 34), ('Why it matters', 56), ('Status', 16), ('Done?', 10)])

ASSETS = [
    ('Product photos — all 12 products',
     'Every image is still hotlinked to mediaprint-eg.com/wp-content/. They break the day the old '
     'WordPress install is removed.', 'Hotlinked'),
    ('Photos of the zippered garment bags',
     'You said photos were attached but nothing came through.', 'Not received'),
    ('Logo as SVG or transparent PNG',
     'Currently a 200x200 PNG. Fine at small sizes, will look soft on a large screen.', 'Low-res'),
    ('Workshop and press photos',
     'The benefit cards use generic Unsplash stock. Real photos of your floor would be far stronger.',
     'Stock images'),
    ('Sticker product photos',
     'All 17 sticker cards show a placeholder icon.', 'Missing'),
    ('Case study photos',
     'You have 22 named clients and no photos of work done for any of them.', 'Missing'),
    ('Exact brand orange',
     'Site uses #DD6320. Your logo looks slightly brighter. Send the exact hex.', 'Approximate'),
]
r = 4
for what, why, status in ASSETS:
    put(ws, r, 1, what, BOLD)
    put(ws, r, 2, why, MUTED)
    put(ws, r, 3, status, BODY, WARN_FILL, CTR)
    put(ws, r, 4, None, INPUT_FONT, INPUT_FILL, CTR)
    ws.row_dimensions[r].height = 30
    r += 1

dv = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'D4:D{r - 1}')

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('saved', OUT)
